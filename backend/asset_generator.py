"""
HoneyNet Dynamic Deception Asset Generator
Generates high-fidelity synthetic company artifacts using openpyxl (.xlsx),
CSV, and structured text, populated with consistent corporate identities from Faker.
Enforces strict filesystem sandbox boundaries to prevent path traversal escapes.
"""
import os
import re
import csv
import logging
from pathlib import Path
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.config import HONEYFS_DIR
from backend.identity_seeder import CompanyIdentity, generate_company_identity

logger = logging.getLogger("honeynet.asset_generator")

# Strict path traversal validator
SAFE_PATH_REGEX = re.compile(r"^[a-zA-Z0-9_\-./]+$")

def sanitize_and_resolve_path(rel_path: str) -> Path:
    """
    Validates that the given relative path stays strictly within the Cowrie honeyfs directory.
    Rejects any path containing '..', null bytes, or shell metacharacters.
    """
    clean_path_str = rel_path.strip().lstrip("/")
    if ".." in clean_path_str or "\x00" in clean_path_str or not SAFE_PATH_REGEX.match(clean_path_str):
        raise ValueError(f"Path traversal security violation: {rel_path}")
    
    full_path = (HONEYFS_DIR / clean_path_str).resolve()
    honeyfs_root = HONEYFS_DIR.resolve()
    
    if not str(full_path).startswith(str(honeyfs_root)):
        raise ValueError(f"Resolved path escapes honeyfs sandbox: {full_path}")
        
    return full_path

def generate_payroll_xlsx(dest_path: Path, company: CompanyIdentity) -> Path:
    """
    Creates an authentic binary Microsoft Excel (.xlsx) payroll workbook using openpyxl.
    Features formatted headers, corporate styling, numeric cells, and auto-adjusted column widths.
    """
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026 Executive Payroll"
    
    # 1. Company Banner Header
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"CONFIDENTIAL — {company.name.upper()} — 2026 CORPORATE PAYROLL"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle / Metadata
    ws["A2"] = f"Tax ID / EIN: {company.tax_id} | Domain: {company.domain} | HQ: {company.headquarters}"
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="555555")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 18

    # 2. Table Column Headers
    headers = [
        "Employee ID", "Full Legal Name", "Department", "Corporate Role",
        "Corporate Email", "Base Salary (USD)", "Target Bonus (USD)", "Direct Deposit Routing"
    ]
    
    ws.append([]) # Empty row 3
    ws.append(headers) # Row 4
    ws.row_dimensions[4].height = 24
    
    header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_border = Border(
        bottom=Side(style="medium", color="1A365D"),
        top=Side(style="thin", color="CCCCCC")
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = header_border

    # 3. Insert Employee Data Rows
    thin_border = Border(
        bottom=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0")
    )
    
    start_row = 5
    for idx, emp in enumerate(company.employees):
        row_num = start_row + idx
        raw_salary = float(emp.salary.replace("$", "").replace(",", ""))
        raw_bonus = float(emp.annual_bonus.replace("$", "").replace(",", ""))
        mock_routing = f"****{1000 + idx}"

        row_data = [
            emp.id, emp.name, emp.department, emp.role,
            emp.email, raw_salary, raw_bonus, f"JPMorgan Chase {mock_routing}"
        ]
        ws.append(row_data)
        ws.row_dimensions[row_num].height = 20
        
        # Format styles
        for col_num in range(1, len(row_data) + 1):
            c = ws.cell(row=row_num, column=col_num)
            c.font = Font(name="Calibri", size=10)
            c.border = thin_border
            if col_num in (6, 7):
                c.number_format = "$#,##0"
                c.alignment = Alignment(horizontal="right", vertical="center")

    # 4. Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(str(dest_path))
    logger.info(f"Generated authentic Excel payroll at {dest_path}")
    return dest_path

def generate_payroll_csv(dest_path: Path, company: CompanyIdentity) -> Path:
    """Generates matching CSV payroll export."""
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    with open(dest_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["# CONFIDENTIAL PAYROLL RECORD", company.name, company.tax_id])
        writer.writerow(["EmployeeID", "FullName", "Department", "Role", "Email", "Salary", "Bonus", "RoutingAccount"])
        for emp in company.employees:
            writer.writerow([emp.id, emp.name, emp.department, emp.role, emp.email, emp.salary, emp.annual_bonus, f"CHASE-MOCK-****{emp.id[-3:]}"])
    return dest_path

def generate_wire_transfers(dest_path: Path, company: CompanyIdentity) -> Path:
    """Generates wire transfer authorization memos."""
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    cfo = next((e for e in company.employees if e.department == "Finance"), company.employees[0])
    ceo = company.employees[0]
    
    content = f"""================================================================================
{company.name.upper()} — CORPORATE TREASURY & WIRE DISBURSEMENT LOG
CONFIDENTIAL // INTERNAL AUDIT ONLY
================================================================================
Company Tax ID: {company.tax_id}
Treasury Bank: Silicon Valley Bank / First Citizens Division
Operating Account: ********-8821
SWIFT Code: SVBUS6SXXX

AUTHORIZED WIRE TRANSACTIONS (Q1 2026):
--------------------------------------------------------------------------------
1. Date: 2026-02-01 | Amount: $284,500.00 USD
   Beneficiary: AWS Enterprise Billing Inc.
   Account: 0984-2234-9912 | Fedwire Reference: FW-2026-00918
   Signoff: {cfo.name} ({cfo.role})

2. Date: 2026-02-14 | Amount: $1,420,000.00 USD
   Beneficiary: {company.name} Global Payroll Escrow (ADP TotalSource)
   Account: 1102-8871-3341 | Fedwire Reference: FW-2026-01042
   Signoff: {ceo.name} ({ceo.role})

3. Date: 2026-02-28 | Amount: $95,000.00 USD
   Beneficiary: Datadog Cloud Security Services LLC
   Account: 4432-1120-7761 | Fedwire Reference: FW-2026-01209
   Signoff: {cfo.name} ({cfo.role})
================================================================================
"""
    dest_path.write_text(content, encoding="utf-8")
    return dest_path

def generate_git_env(dest_path: Path, company: CompanyIdentity) -> Path:
    """Generates production .env template containing seeded corporate secrets."""
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    clean_domain_name = company.domain.split(".")[0]
    devops_lead = next((e for e in company.employees if e.department == "DevOps"), company.employees[1])
    
    content = f"""# ==============================================================================
# {company.name.upper()} PRODUCTION ENVIRONMENT SECRETS (.env)
# MAINTAINER: {devops_lead.name} <{devops_lead.email}>
# WARNING: STRICTLY CONFIDENTIAL - DO NOT COMMIT TO VERSION CONTROL
# ==============================================================================

NODE_ENV=production
APP_ENV=prod
PORT=8080
DOMAIN={company.domain}
SECRET_KEY=canary_sec_key_{clean_domain_name}_99a81bc3e78f

# Database Credentials
DATABASE_URL=postgresql://{clean_domain_name}_admin:canary_db_pass_2026@db-master.{company.domain}:5432/{clean_domain_name}_prod
REDIS_URL=redis://:{clean_domain_name}_cache_token@redis-cluster.{company.domain}:6379/0

# Stripe Payment Gateway (CANARY TEST TOKENS)
STRIPE_PUBLISHABLE_KEY=pk_test_canary_51Nz83JKe82910_{clean_domain_name}
STRIPE_SECRET_KEY=sk_test_canary_mock_token_9918237_{clean_domain_name}
STRIPE_WEBHOOK_SECRET=whsec_mock_canary_sig_test_2026

# Internal SSO & OAuth
GOOGLE_CLIENT_ID=canary-app-881923.{clean_domain_name}.apps.googleusercontent.com
GOOGLE_CLIENT_SECRET=GOCSPX-canary-mock-secret-key-2026
JWT_SIGNING_SECRET=canary_jwt_sign_token_88921_{clean_domain_name}
"""
    dest_path.write_text(content, encoding="utf-8")
    return dest_path

def generate_dynamic_deception_assets(
    category: str,
    session_id: str,
    company: Optional[CompanyIdentity] = None
) -> List[dict]:
    """
    Generates dynamic honeytoken artifacts on-demand into Cowrie's fake fs sandbox.
    Uses seeded company identity to maintain absolute cross-file consistency.
    """
    if not company:
        company = generate_company_identity(session_id)
        
    created_assets = []
    
    try:
        if category in ("finance", "payroll", "salary"):
            # 1. Authentic Excel Payroll
            xlsx_path = sanitize_and_resolve_path("home/phil/finance/Payroll_2026_Confidential.xlsx")
            generate_payroll_xlsx(xlsx_path, company)
            created_assets.append({
                "type": "binary_xlsx",
                "path": "/home/phil/finance/Payroll_2026_Confidential.xlsx",
                "content_ref": f"Excel Workbook with {len(company.employees)} employees ({company.name})",
                "category": "finance"
            })

            # 2. CSV Payroll
            csv_path = sanitize_and_resolve_path("home/phil/finance/Payroll_2026_Confidential.csv")
            generate_payroll_csv(csv_path, company)
            created_assets.append({
                "type": "csv_table",
                "path": "/home/phil/finance/Payroll_2026_Confidential.csv",
                "content_ref": f"CSV Payroll Export for {company.name}",
                "category": "finance"
            })

            # 3. Wire Transfers
            wire_path = sanitize_and_resolve_path("home/phil/finance/Wire_Transfer_Authorizations.txt")
            generate_wire_transfers(wire_path, company)
            created_assets.append({
                "type": "treasury_memo",
                "path": "/home/phil/finance/Wire_Transfer_Authorizations.txt",
                "content_ref": f"Corporate Wire Disbursement Log ({company.tax_id})",
                "category": "finance"
            })

        elif category in ("git", "credentials", "secret"):
            env_path = sanitize_and_resolve_path("home/phil/git/.env")
            generate_git_env(env_path, company)
            created_assets.append({
                "type": "env_credentials",
                "path": "/home/phil/git/.env",
                "content_ref": f"Production .env with {company.domain} canary tokens",
                "category": "git"
            })

    except Exception as e:
        logger.error(f"Error generating deception assets for session {session_id}: {e}")

    return created_assets
